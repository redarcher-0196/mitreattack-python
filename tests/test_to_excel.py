"""Tests for ATT&CK to Excel export behavior."""

from dataclasses import dataclass
from pathlib import Path

import pandas as pd
import pytest
import stix2
from openpyxl import load_workbook

from mitreattack.attackToExcel import attackToExcel


def _object_data(object_type: str, rows=None, citations=None):
    return {
        object_type: pd.DataFrame(rows or [{"ID": f"{object_type}-1", "name": f"{object_type} one"}]),
        "citations": pd.DataFrame(citations or [{"reference": "alpha", "citation": "Alpha Citation"}]),
    }


def _sheet_names(path: Path):
    workbook = load_workbook(path, read_only=True)
    try:
        return workbook.sheetnames
    finally:
        workbook.close()


def _sheet_rows(path: Path, sheet_name: str):
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return [tuple(cell for cell in row) for row in workbook[sheet_name].iter_rows(values_only=True)]
    finally:
        workbook.close()


def _write_stix_files(stix_base_dir: Path, domains):
    stix_base_dir.mkdir(parents=True, exist_ok=True)
    for domain in domains:
        (stix_base_dir / f"{domain}.json").write_text("{}", encoding="utf-8")


def _release_output_dir(output_dir: Path, version: str | None = None) -> Path:
    return output_dir / version if version else output_dir


def _staging_output_dir(output_dir: Path, version: str | None = None) -> Path:
    return _release_output_dir(output_dir, version) / "tmp" / "staged-excel-files"


def _domain_export_dir(
    output_dir: Path, domain: str, version: str | None = None, *, versioned_output_dir: bool = False
):
    release_output_dir = _release_output_dir(output_dir, version)
    domain_dir_name = f"{domain}-{version}" if version and versioned_output_dir else domain
    return release_output_dir / domain_dir_name


def _write_existing_release_excel(
    output_dir: Path,
    domain: str,
    version: str | None = None,
    *,
    versioned_output_dir: bool = False,
):
    domain_dir = _domain_export_dir(
        output_dir,
        domain,
        version,
        versioned_output_dir=versioned_output_dir,
    )
    domain_dir.mkdir(parents=True, exist_ok=True)
    workbook_name = f"{domain}-{version}.xlsx" if version else f"{domain}.xlsx"
    existing_file = domain_dir / workbook_name
    existing_file.write_text("existing", encoding="utf-8")
    return existing_file


def _make_fake_release_export(*, calls=None, log_marker=None):
    def fake_export(**kwargs):
        if calls is not None:
            calls.append(kwargs)
        if log_marker is not None:
            log_marker()

        output_dir = Path(kwargs["output_dir"])
        domain_version_string = f"{kwargs['domain']}-{kwargs['version']}" if kwargs["version"] else kwargs["domain"]
        versioned_dir = output_dir / domain_version_string
        versioned_dir.mkdir(parents=True)
        (versioned_dir / f"{domain_version_string}.xlsx").write_text("new", encoding="utf-8")

    return fake_export


@dataclass
class FakeMergeRange:
    """Small stand-in for matrix merge range objects."""

    data: str = "Merged Header"
    leftCol: int = 1
    rightCol: int = 2
    format: dict | None = None

    def to_excel_format(self):
        """Return the Excel cell range used by the matrix writer."""
        return "A4:B4"


def test_export_with_memstore_uses_current_dataframe_builder(
    monkeypatch,
    tmp_path: Path,
    attack_memstore_factory,
    sample_technique_object,
):
    """Current exports should build v18+ dataframes and pass them to write_excel."""
    mem_store = attack_memstore_factory([sample_technique_object])
    dataframes = {"techniques": _object_data("techniques")}
    calls = {}

    def fake_build_dataframes(**kwargs):
        calls["build_dataframes"] = kwargs
        return dataframes

    def fake_build_dataframes_pre_v18(**kwargs):
        calls["build_dataframes_pre_v18"] = kwargs
        return {}

    def fake_write_excel(**kwargs):
        calls["write_excel"] = kwargs

    monkeypatch.setattr(attackToExcel, "build_dataframes", fake_build_dataframes)
    monkeypatch.setattr(attackToExcel, "build_dataframes_pre_v18", fake_build_dataframes_pre_v18)
    monkeypatch.setattr(attackToExcel, "write_excel", fake_write_excel)

    attackToExcel.export(domain="enterprise-attack", output_dir=str(tmp_path), mem_store=mem_store)

    assert calls["build_dataframes"] == {"src": mem_store, "domain": "enterprise-attack"}
    assert "build_dataframes_pre_v18" not in calls
    assert calls["write_excel"] == {
        "dataframes": dataframes,
        "domain": "enterprise-attack",
        "src": mem_store,
        "version": None,
        "output_dir": str(tmp_path),
        "overwrite": False,
        "log_written_files": True,
    }


def test_export_with_pre_v18_version_uses_legacy_dataframe_builder(
    monkeypatch,
    tmp_path: Path,
    attack_memstore_factory,
    sample_technique_object,
):
    """Pre-v18 exports should use the legacy dataframe builder."""
    mem_store = attack_memstore_factory([sample_technique_object])
    dataframes = {"techniques": _object_data("techniques")}
    calls = {}

    def fake_build_dataframes(**kwargs):
        calls["build_dataframes"] = kwargs
        return {}

    def fake_build_dataframes_pre_v18(**kwargs):
        calls["build_dataframes_pre_v18"] = kwargs
        return dataframes

    def fake_write_excel(**kwargs):
        calls["write_excel"] = kwargs

    monkeypatch.setattr(attackToExcel, "build_dataframes", fake_build_dataframes)
    monkeypatch.setattr(attackToExcel, "build_dataframes_pre_v18", fake_build_dataframes_pre_v18)
    monkeypatch.setattr(attackToExcel, "write_excel", fake_write_excel)

    attackToExcel.export(domain="enterprise-attack", version="v17.0", output_dir=str(tmp_path), mem_store=mem_store)

    assert calls["build_dataframes_pre_v18"] == {"src": mem_store, "domain": "enterprise-attack"}
    assert "build_dataframes" not in calls
    assert calls["write_excel"]["version"] == "v17.0"
    assert calls["write_excel"]["dataframes"] is dataframes
    assert calls["write_excel"]["overwrite"] is False
    assert calls["write_excel"]["log_written_files"] is True


def test_export_refuses_existing_excel_file_before_building_dataframes(
    monkeypatch,
    tmp_path: Path,
    attack_memstore_factory,
    sample_technique_object,
):
    """Export should refuse existing Excel files before building or writing dataframes."""
    output_directory = tmp_path / "enterprise-attack"
    output_directory.mkdir()
    existing_file = output_directory / "enterprise-attack-techniques.xlsx"
    existing_file.write_text("existing", encoding="utf-8")
    mem_store = attack_memstore_factory([sample_technique_object])
    calls = []

    def fake_build_dataframes(**kwargs):
        calls.append(kwargs)
        return {"techniques": _object_data("techniques")}

    monkeypatch.setattr(attackToExcel, "build_dataframes", fake_build_dataframes)

    with pytest.raises(FileExistsError, match="Refusing to overwrite existing Excel file"):
        attackToExcel.export(domain="enterprise-attack", output_dir=str(tmp_path), mem_store=mem_store)

    assert calls == []


def test_export_overwrite_allows_existing_excel_file(
    monkeypatch,
    tmp_path: Path,
    attack_memstore_factory,
    sample_technique_object,
):
    """Export should build and write when overwrite is explicitly requested."""
    output_directory = tmp_path / "enterprise-attack"
    output_directory.mkdir()
    (output_directory / "enterprise-attack-techniques.xlsx").write_text("existing", encoding="utf-8")
    mem_store = attack_memstore_factory([sample_technique_object])
    calls = {}

    def fake_build_dataframes(**kwargs):
        return {"techniques": _object_data("techniques")}

    def fake_write_excel(**kwargs):
        calls["write_excel"] = kwargs

    monkeypatch.setattr(attackToExcel, "build_dataframes", fake_build_dataframes)
    monkeypatch.setattr(attackToExcel, "write_excel", fake_write_excel)

    attackToExcel.export(domain="enterprise-attack", output_dir=str(tmp_path), mem_store=mem_store, overwrite=True)

    assert calls["write_excel"]["overwrite"] is True


def test_export_rejects_multiple_stix_sources(attack_memstore_factory, sample_technique_object):
    """Export should reject ambiguous STIX source inputs before building dataframes."""
    mem_store = attack_memstore_factory([sample_technique_object])

    with pytest.raises(TypeError, match="Exactly zero or one"):
        attackToExcel.export(remote="http://localhost:3000", stix_file="bundle.json", mem_store=mem_store)


def test_normalize_attack_version_adds_missing_prefix():
    """ATT&CK release versions should be normalized to release directory names."""
    assert attackToExcel.normalize_attack_version("19.0") == "v19.0"
    assert attackToExcel.normalize_attack_version("v19.0") == "v19.0"


def test_export_release_uses_existing_local_stix_files(tmp_path: Path, monkeypatch):
    """Release export should use existing local STIX files without downloading."""
    stix_base_dir = tmp_path / "attack-releases" / "stix-2.0" / "v19.0"
    _write_stix_files(stix_base_dir, ["enterprise-attack", "mobile-attack"])

    calls = {}

    def fake_download_domains(**kwargs):
        calls.setdefault("downloads", []).append(kwargs)

    def fake_export(**kwargs):
        calls.setdefault("exports", []).append(kwargs)

    monkeypatch.setattr(attackToExcel, "download_domains", fake_download_domains)
    monkeypatch.setattr(attackToExcel, "export", fake_export)

    attackToExcel.export_release(
        version="19.0",
        stix_base_dir=str(stix_base_dir),
        output_dir=str(tmp_path / "output"),
        domains=["enterprise-attack", "mobile-attack"],
    )

    assert "downloads" not in calls
    assert [call["domain"] for call in calls["exports"]] == ["enterprise-attack", "mobile-attack"]
    assert calls["exports"][0]["stix_file"] == str(stix_base_dir / "enterprise-attack.json")
    assert calls["exports"][0]["version"] == "v19.0"
    assert calls["exports"][0]["output_dir"] == str(_staging_output_dir(tmp_path / "output", "v19.0"))
    assert calls["exports"][0]["overwrite"] is False


def test_export_release_with_explicit_local_stix_base_dir_without_version_is_unversioned(tmp_path: Path, monkeypatch):
    """Explicit local STIX bundle directories should not be labelled as ATT&CK releases unless a version is given."""
    stix_base_dir = tmp_path / "attack-releases" / "stix-2.0" / "attackwb"
    _write_stix_files(stix_base_dir, ["enterprise-attack", "mobile-attack"])

    calls = {}

    def fake_export(**kwargs):
        calls.setdefault("exports", []).append(kwargs)

    monkeypatch.setattr(attackToExcel, "export", fake_export)

    attackToExcel.export_release(
        stix_base_dir=str(stix_base_dir),
        output_dir=str(tmp_path / "output" / "attackwb"),
        domains=["enterprise-attack", "mobile-attack"],
    )

    assert [call["domain"] for call in calls["exports"]] == ["enterprise-attack", "mobile-attack"]
    assert calls["exports"][0]["version"] is None
    assert calls["exports"][0]["output_dir"] == str(_staging_output_dir(tmp_path / "output" / "attackwb"))
    assert calls["exports"][0]["overwrite"] is False
    assert calls["exports"][1]["version"] is None
    assert calls["exports"][1]["output_dir"] == str(_staging_output_dir(tmp_path / "output" / "attackwb"))


def test_export_release_downloads_only_missing_domains_to_temporary_directory(tmp_path: Path, monkeypatch):
    """Missing release STIX files should be downloaded per missing domain into a temporary tree."""
    stix_base_dir = tmp_path / "attack-releases" / "stix-2.0" / "v19.0"
    stix_base_dir.mkdir(parents=True)
    (stix_base_dir / "enterprise-attack.json").write_text("{}", encoding="utf-8")
    calls = {}

    def fake_download_domains(**kwargs):
        calls["download"] = kwargs
        release_dir = Path(kwargs["download_dir"]) / "v19.0"
        release_dir.mkdir(parents=True)
        for domain in kwargs["domains"]:
            (release_dir / f"{domain}-attack.json").write_text("{}", encoding="utf-8")

    def fake_export(**kwargs):
        calls.setdefault("exports", []).append(kwargs)
        assert Path(kwargs["stix_file"]).exists()

    monkeypatch.setattr(attackToExcel, "download_domains", fake_download_domains)
    monkeypatch.setattr(attackToExcel, "export", fake_export)

    attackToExcel.export_release(
        version="v19.0",
        stix_base_dir=str(stix_base_dir),
        output_dir=str(tmp_path / "output"),
        domains=["enterprise-attack", "mobile-attack", "ics-attack"],
    )

    assert calls["download"]["domains"] == ["mobile", "ics"]
    assert calls["download"]["all_versions"] is False
    assert calls["download"]["stix_version"] == "2.0"
    assert calls["download"]["attack_versions"] == ["19.0"]
    assert calls["exports"][0]["stix_file"] == str(stix_base_dir / "enterprise-attack.json")
    assert calls["exports"][1]["stix_file"].endswith("stix-2.0/v19.0/mobile-attack.json")
    assert calls["exports"][2]["stix_file"].endswith("stix-2.0/v19.0/ics-attack.json")
    assert not Path(calls["exports"][1]["stix_file"]).exists()


def test_export_release_moves_versioned_outputs_to_domain_directory(tmp_path: Path, monkeypatch):
    """Default release export should flatten domain-version folders into domain folders."""
    stix_base_dir = tmp_path / "stix"
    _write_stix_files(stix_base_dir, ["enterprise-attack"])

    monkeypatch.setattr(attackToExcel, "export", _make_fake_release_export())

    attackToExcel.export_release(
        version="v19.0",
        stix_base_dir=str(stix_base_dir),
        output_dir=str(tmp_path / "output"),
        domains=["enterprise-attack"],
    )

    assert not (tmp_path / "output" / "v19.0" / "enterprise-attack-v19.0").exists()
    assert (tmp_path / "output" / "v19.0" / "enterprise-attack" / "enterprise-attack-v19.0.xlsx").exists()


@pytest.mark.parametrize("versioned_output_dir", [False, True])
def test_export_release_refuses_existing_excel_file_before_exporting(
    tmp_path: Path,
    monkeypatch,
    versioned_output_dir: bool,
):
    """Release export should check all selected domain output folders before exporting anything."""
    output_dir = tmp_path / "output"
    _write_existing_release_excel(
        output_dir,
        "mobile-attack",
        "v19.0",
        versioned_output_dir=versioned_output_dir,
    )

    stix_base_dir = tmp_path / "stix"
    _write_stix_files(stix_base_dir, ["enterprise-attack", "mobile-attack"])

    calls = []

    def fake_export(**kwargs):
        calls.append(kwargs)

    monkeypatch.setattr(attackToExcel, "export", fake_export)

    with pytest.raises(FileExistsError, match="Refusing to overwrite existing Excel file"):
        attackToExcel.export_release(
            version="v19.0",
            stix_base_dir=str(stix_base_dir),
            output_dir=str(output_dir),
            domains=["enterprise-attack", "mobile-attack"],
            versioned_output_dir=versioned_output_dir,
        )

    assert calls == []


@pytest.mark.parametrize("versioned_output_dir", [False, True])
def test_export_release_overwrite_allows_existing_excel_file(
    tmp_path: Path,
    monkeypatch,
    versioned_output_dir: bool,
):
    """Release export should proceed when overwrite is explicitly requested."""
    output_dir = tmp_path / "output"
    _write_existing_release_excel(
        output_dir,
        "mobile-attack",
        "v19.0",
        versioned_output_dir=versioned_output_dir,
    )

    stix_base_dir = tmp_path / "stix"
    _write_stix_files(stix_base_dir, ["mobile-attack"])

    calls = []

    def fake_export(**kwargs):
        calls.append(kwargs)

    monkeypatch.setattr(attackToExcel, "export", fake_export)

    attackToExcel.export_release(
        version="v19.0",
        stix_base_dir=str(stix_base_dir),
        output_dir=str(output_dir),
        domains=["mobile-attack"],
        versioned_output_dir=versioned_output_dir,
        overwrite=True,
    )

    assert calls[0]["overwrite"] is True
    assert calls[0]["log_written_files"] is versioned_output_dir


def test_export_release_logs_each_flattened_file_overwrite(tmp_path: Path, monkeypatch):
    """Release export should log target overwrites before staging files."""
    output_dir = tmp_path / "output"
    existing_file = _write_existing_release_excel(output_dir, "enterprise-attack", "v19.0")

    stix_base_dir = tmp_path / "stix"
    _write_stix_files(stix_base_dir, ["enterprise-attack"])
    log_messages = []

    monkeypatch.setattr(
        attackToExcel,
        "export",
        _make_fake_release_export(log_marker=lambda: log_messages.append("fake export started")),
    )
    monkeypatch.setattr(attackToExcel.logger, "info", log_messages.append)

    attackToExcel.export_release(
        version="v19.0",
        stix_base_dir=str(stix_base_dir),
        output_dir=str(output_dir),
        domains=["enterprise-attack"],
        overwrite=True,
    )

    overwrite_message = f"Overwriting existing Excel file: {existing_file}"
    assert "Existing Excel files will be overwritten:" in log_messages
    assert overwrite_message in log_messages
    assert log_messages.index(overwrite_message) < log_messages.index("fake export started")


def test_export_release_logs_staging_and_final_move_directories(tmp_path: Path, monkeypatch):
    """Default release export should identify staged writes and final output moves."""
    output_dir = tmp_path / "output"
    stix_base_dir = tmp_path / "stix"
    _write_stix_files(stix_base_dir, ["enterprise-attack"])
    log_messages = []

    monkeypatch.setattr(attackToExcel, "export", _make_fake_release_export())
    monkeypatch.setattr(attackToExcel.logger, "info", log_messages.append)

    attackToExcel.export_release(
        version="v19.0",
        stix_base_dir=str(stix_base_dir),
        output_dir=str(output_dir),
        domains=["enterprise-attack"],
        overwrite=True,
    )

    staging_dir = _staging_output_dir(output_dir, "v19.0") / "enterprise-attack-v19.0"
    assert f"Writing staged Excel files for enterprise-attack to {staging_dir}" in log_messages
    assert f"Moving staged Excel files for enterprise-attack to {output_dir / 'v19.0' / 'enterprise-attack'}" in (
        log_messages
    )


def test_export_release_logs_written_files_after_final_move(tmp_path: Path, monkeypatch):
    """Release export should report final output paths after staged files are moved."""
    output_dir = tmp_path / "output"
    stix_base_dir = tmp_path / "stix"
    _write_stix_files(stix_base_dir, ["enterprise-attack"])
    log_messages = []

    monkeypatch.setattr(attackToExcel, "export", _make_fake_release_export())
    monkeypatch.setattr(attackToExcel.logger, "info", log_messages.append)

    attackToExcel.export_release(
        version="v19.0",
        stix_base_dir=str(stix_base_dir),
        output_dir=str(output_dir),
        domains=["enterprise-attack"],
    )

    staged_file = _staging_output_dir(output_dir, "v19.0") / "enterprise-attack-v19.0" / "enterprise-attack-v19.0.xlsx"
    final_file = output_dir / "v19.0" / "enterprise-attack" / "enterprise-attack-v19.0.xlsx"
    assert f"Excel file written: {final_file}" in log_messages
    assert f"Excel file written: {staged_file}" not in log_messages


def test_export_release_uses_tmp_staging_directory(tmp_path: Path, monkeypatch):
    """Default release export should stage workbooks under tmp/staged-excel-files."""
    output_dir = tmp_path / "output"
    stix_base_dir = tmp_path / "stix"
    _write_stix_files(stix_base_dir, ["enterprise-attack"])
    calls = []

    monkeypatch.setattr(attackToExcel, "export", _make_fake_release_export(calls=calls))

    attackToExcel.export_release(
        version="v19.0",
        stix_base_dir=str(stix_base_dir),
        output_dir=str(output_dir),
        domains=["enterprise-attack"],
    )

    staging_output_dir = _staging_output_dir(output_dir, "v19.0")
    assert calls[0]["output_dir"] == str(staging_output_dir)
    assert not (staging_output_dir / "enterprise-attack-v19.0").exists()
    assert (output_dir / "v19.0" / "enterprise-attack" / "enterprise-attack-v19.0.xlsx").exists()


def test_export_release_moves_unversioned_staged_outputs_to_domain_directory(tmp_path: Path, monkeypatch):
    """Unversioned release exports should move staged domain folders to final domain folders."""
    output_dir = tmp_path / "output" / "attackwb"
    stix_base_dir = tmp_path / "stix"
    _write_stix_files(stix_base_dir, ["enterprise-attack"])

    monkeypatch.setattr(attackToExcel, "export", _make_fake_release_export())

    attackToExcel.export_release(
        stix_base_dir=str(stix_base_dir),
        output_dir=str(output_dir),
        domains=["enterprise-attack"],
    )

    assert not (_staging_output_dir(output_dir) / "enterprise-attack").exists()
    assert (output_dir / "enterprise-attack" / "enterprise-attack.xlsx").exists()


def test_export_release_rejects_invalid_domain():
    """Release export should validate selected ATT&CK domains."""
    with pytest.raises(ValueError, match="Invalid ATT&CK domain"):
        attackToExcel.export_release(domains=["pre-attack"])


def test_write_excel_creates_expected_workbooks(monkeypatch, tmp_path: Path, attack_memstore_factory):
    """write_excel should create a master workbook and object-specific workbooks."""
    monkeypatch.setattr(attackToExcel.stixToDf, "detectionStrategiesAnalyticsLogSourcesDf", lambda src: pd.DataFrame())
    mem_store = attack_memstore_factory([])
    dataframes = {
        "techniques": _object_data("techniques"),
        "groups": _object_data("groups"),
    }

    written_files = attackToExcel.write_excel(
        dataframes=dataframes,
        domain="enterprise-attack",
        src=mem_store,
        output_dir=str(tmp_path),
    )

    output_folder = tmp_path / "enterprise-attack"
    assert set(map(Path, written_files)) == {
        output_folder / "enterprise-attack-techniques.xlsx",
        output_folder / "enterprise-attack-groups.xlsx",
        output_folder / "enterprise-attack.xlsx",
    }
    assert _sheet_names(output_folder / "enterprise-attack.xlsx") == ["techniques", "groups", "citations"]


def test_write_excel_logs_each_file_overwrite(monkeypatch, tmp_path: Path, attack_memstore_factory):
    """write_excel should log each generated workbook replaced by overwrite."""
    monkeypatch.setattr(attackToExcel.stixToDf, "detectionStrategiesAnalyticsLogSourcesDf", lambda src: pd.DataFrame())
    mem_store = attack_memstore_factory([])
    output_folder = tmp_path / "enterprise-attack"
    output_folder.mkdir()
    existing_files = [
        output_folder / "enterprise-attack.xlsx",
        output_folder / "enterprise-attack-techniques.xlsx",
        output_folder / "enterprise-attack-groups.xlsx",
    ]
    for existing_file in existing_files:
        existing_file.write_text("existing", encoding="utf-8")

    log_messages = []
    monkeypatch.setattr(attackToExcel.logger, "info", log_messages.append)

    attackToExcel.write_excel(
        dataframes={
            "techniques": _object_data("techniques"),
            "groups": _object_data("groups"),
        },
        domain="enterprise-attack",
        src=mem_store,
        output_dir=str(tmp_path),
        overwrite=True,
    )

    for existing_file in existing_files:
        assert f"Overwriting existing Excel file: {existing_file}" in log_messages


def test_write_excel_logs_files_written(monkeypatch, tmp_path: Path, attack_memstore_factory):
    """write_excel should use wording that applies to created and overwritten workbooks."""
    monkeypatch.setattr(attackToExcel.stixToDf, "detectionStrategiesAnalyticsLogSourcesDf", lambda src: pd.DataFrame())
    mem_store = attack_memstore_factory([])
    log_messages = []

    monkeypatch.setattr(attackToExcel.logger, "info", log_messages.append)

    attackToExcel.write_excel(
        dataframes={"techniques": _object_data("techniques")},
        domain="enterprise-attack",
        src=mem_store,
        output_dir=str(tmp_path),
    )

    output_folder = tmp_path / "enterprise-attack"
    assert f"Excel file written: {output_folder / 'enterprise-attack-techniques.xlsx'}" in log_messages
    assert f"Excel file written: {output_folder / 'enterprise-attack.xlsx'}" in log_messages
    assert not any("Excel file created:" in message for message in log_messages)


def test_write_excel_can_suppress_written_file_logs(monkeypatch, tmp_path: Path, attack_memstore_factory):
    """Release staging can suppress staged-path written logs."""
    monkeypatch.setattr(attackToExcel.stixToDf, "detectionStrategiesAnalyticsLogSourcesDf", lambda src: pd.DataFrame())
    mem_store = attack_memstore_factory([])
    log_messages = []

    monkeypatch.setattr(attackToExcel.logger, "info", log_messages.append)

    attackToExcel.write_excel(
        dataframes={"techniques": _object_data("techniques")},
        domain="enterprise-attack",
        src=mem_store,
        output_dir=str(tmp_path),
        log_written_files=False,
    )

    assert not any("Excel file written:" in message for message in log_messages)


def test_write_excel_skips_empty_object_data(monkeypatch, tmp_path: Path, attack_memstore_factory):
    """Empty object data should not produce an object workbook."""
    monkeypatch.setattr(attackToExcel.stixToDf, "detectionStrategiesAnalyticsLogSourcesDf", lambda src: pd.DataFrame())
    mem_store = attack_memstore_factory([])
    dataframes = {
        "techniques": {},
        "groups": _object_data("groups"),
    }

    attackToExcel.write_excel(
        dataframes=dataframes,
        domain="enterprise-attack",
        src=mem_store,
        output_dir=str(tmp_path),
    )

    output_folder = tmp_path / "enterprise-attack"
    assert not (output_folder / "enterprise-attack-techniques.xlsx").exists()
    assert (output_folder / "enterprise-attack-groups.xlsx").exists()


def test_write_excel_dedupes_and_sorts_citations(monkeypatch, tmp_path: Path, attack_memstore_factory):
    """Master citations should be deduped by reference and sorted by reference."""
    monkeypatch.setattr(attackToExcel.stixToDf, "detectionStrategiesAnalyticsLogSourcesDf", lambda src: pd.DataFrame())
    mem_store = attack_memstore_factory([])
    dataframes = {
        "techniques": _object_data(
            "techniques",
            citations=[
                {"reference": "zeta", "citation": "Zeta Citation"},
                {"reference": "alpha", "citation": "Alpha Citation"},
            ],
        ),
        "groups": _object_data("groups", citations=[{"reference": "alpha", "citation": "Duplicate Alpha"}]),
    }

    attackToExcel.write_excel(
        dataframes=dataframes,
        domain="enterprise-attack",
        src=mem_store,
        output_dir=str(tmp_path),
    )

    rows = _sheet_rows(tmp_path / "enterprise-attack" / "enterprise-attack.xlsx", "citations")
    assert rows == [
        ("reference", "citation"),
        ("alpha", "Alpha Citation"),
        ("zeta", "Zeta Citation"),
    ]


def test_write_excel_adds_defensive_mappings_sheet(monkeypatch, tmp_path: Path, attack_memstore_factory):
    """Defensive mappings should be added to relevant object workbooks and the master workbook."""
    defensive_mappings = pd.DataFrame([{"analytic_id": "AN0001", "analytic_name": "Analytic One"}])
    monkeypatch.setattr(
        attackToExcel.stixToDf,
        "detectionStrategiesAnalyticsLogSourcesDf",
        lambda src: defensive_mappings,
    )
    mem_store = attack_memstore_factory([])
    dataframes = {
        "detectionstrategies": _object_data("detectionstrategies"),
        "analytics": _object_data("analytics"),
        "datacomponents": _object_data("datacomponents"),
    }

    attackToExcel.write_excel(
        dataframes=dataframes,
        domain="enterprise-attack",
        src=mem_store,
        output_dir=str(tmp_path),
    )

    output_folder = tmp_path / "enterprise-attack"
    assert "defensive mappings" in _sheet_names(output_folder / "enterprise-attack.xlsx")
    assert "defensive mappings" in _sheet_names(output_folder / "enterprise-attack-detectionstrategies.xlsx")
    assert "defensive mappings" in _sheet_names(output_folder / "enterprise-attack-analytics.xlsx")
    assert "defensive mappings" in _sheet_names(output_folder / "enterprise-attack-datacomponents.xlsx")


def test_write_excel_sanitizes_matrix_sheet_names_and_applies_merges(
    monkeypatch,
    tmp_path: Path,
    attack_memstore_factory,
):
    """Matrix sheet names should be Excel-safe and merged headers should be applied."""
    monkeypatch.setattr(attackToExcel.stixToDf, "detectionStrategiesAnalyticsLogSourcesDf", lambda src: pd.DataFrame())
    mem_store = attack_memstore_factory([])
    matrix_name = "Bad/Matrix:Name?With[Chars]*AndAVeryLongSuffix"
    second_matrix_name = "Second/Matrix:Name?With[Chars]*AndAVeryLongSuffix"
    dataframes = {
        "techniques": _object_data("techniques"),
        "matrices": (
            [
                {
                    "name": matrix_name,
                    "matrix": pd.DataFrame([{"tactic one": "Technique", "tactic two": "Other"}]),
                    "columns": 2,
                    "merge": [
                        FakeMergeRange(format={"name": "tacticHeader", "format": {"bold": True}}),
                    ],
                }
            ],
            [
                {
                    "name": second_matrix_name,
                    "matrix": pd.DataFrame([{"tactic one": "Subtechnique", "tactic two": "Other"}]),
                    "columns": 2,
                    "merge": [],
                }
            ],
        ),
    }

    attackToExcel.write_excel(
        dataframes=dataframes,
        domain="enterprise-attack",
        src=mem_store,
        output_dir=str(tmp_path),
    )

    output_folder = tmp_path / "enterprise-attack"
    matrix_sheets = _sheet_names(output_folder / "enterprise-attack-matrices.xlsx")
    assert len(matrix_sheets) == 2
    assert all(sheet.endswith("...") for sheet in matrix_sheets)
    assert all(not any(character in sheet for character in attackToExcel.INVALID_CHARACTERS) for sheet in matrix_sheets)
    assert matrix_sheets[0] in _sheet_names(output_folder / "enterprise-attack.xlsx")


@pytest.mark.integration
@pytest.mark.slow
def test_enterprise_latest_smoke(tmp_path: Path, memstore_enterprise_latest: stix2.MemoryStore):
    """A full Enterprise export smoke test is available for explicit integration runs."""
    domain = "enterprise-attack"

    attackToExcel.export(domain=domain, output_dir=str(tmp_path), mem_store=memstore_enterprise_latest)

    excel_folder = tmp_path / domain
    assert (excel_folder / f"{domain}.xlsx").exists()
    assert (excel_folder / f"{domain}-techniques.xlsx").exists()
